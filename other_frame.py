import customtkinter as ctk
from tkcalendar import Calendar
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime


###########################
def create_other_frame(parent):
    other_frame = ctk.CTkFrame(parent, fg_color='#FDEBD0')
    other_frame.place(relwidth=1, relheight=1)

    # ✅ HEADING LABEL
    title_label = ctk.CTkLabel(
        other_frame,
        text="OTHER COLLECTION",
        font=("Arial", 22, "bold"),
        text_color="black"
    )
    title_label.place(relx=0.5, y=10, anchor="n")

    fields = [
        "Date",
        "FTTH No",
        "Name",
        "Contact No",
        "Bill Amount",
        "Cash Received",
        "Balance",
        "Cash With",
        "Collection Type",
        "Remarks"
    ]
    entries = {}
    # ---------- FTTH VALIDATION ----------
    def validate_ftth(new_value):
        if new_value == "":
            return True
        # Only digits allowed
        if not new_value.isdigit():
            return False
        # Maximum 10 digits
        if len(new_value) > 10:
            return False
        # First digit must be 4 only
        if len(new_value) == 1 and new_value != "4":
            return False
        # Extra safety (if pasted full number)
        if len(new_value) > 1 and new_value[0] != "4":
            return False
        # Auto move to Name field after 10 digits
        if len(new_value) == 10:
            other_frame.after(10, lambda: entries["Name"].focus())
        return True

    # ---------- CONTACT VALIDATION ----------
    def validate_contact(new_value):
        if new_value == "":
            return True
        # Only digits allowed
        if not new_value.isdigit():
            return False
        # Maximum 10 digits
        if len(new_value) > 10:
            return False
        # First digit must be 6,7,8,9 only
        if len(new_value) == 1 and new_value not in ["6", "7", "8", "9"]:
            return False
        # Extra safety (if pasted full number)
        if len(new_value) > 1 and new_value[0] not in ["6", "7", "8", "9"]:
            return False
        # Auto move to Bill Amount after 10 digits
        if len(new_value) == 10:
            other_frame.after(10, lambda: entries["Bill Amount"].focus())
        return True

    # ---------- AMOUNT VALIDATION ----------
    def validate_amount(new_value):
        if new_value == "":
            return True
        try:
            float(new_value)
            return True
        except ValueError:
            return False

    vcmd_ftth = (other_frame.register(validate_ftth), "%P")
    vcmd_contact = (other_frame.register(validate_contact), "%P")
    vcmd_amount = (other_frame.register(validate_amount), "%P")

    y_pos = 60
    for field in fields:
        lbl = ctk.CTkLabel(
            other_frame, text=field,
            font=("Arial", 15),
            text_color="black"
        )
        lbl.place(x=20, y=y_pos)
        ############################ Entry Creation Loop
        if field == "Date":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="lightgrey",
                text_color="black",
            )
            # Auto-fill today's date
            widget.insert(0, datetime.now().strftime("%d-%m-%Y"))
            widget.configure(state="readonly")


        elif field == "FTTH No":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="white",
                text_color="black",
                validate="key",
                validatecommand=vcmd_ftth
            )

        elif field == "Contact No":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="white",
                text_color="black",
                validate="key",
                validatecommand=vcmd_contact
            )

        elif field == "Bill Amount":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="white",
                text_color="black",
                validate="key",
                validatecommand=vcmd_amount
            )

        elif field == "Cash Received":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="white",
                text_color="black",
                validate="key",
                validatecommand=vcmd_amount
            )

        elif field == "Balance":
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="lightgrey",
                text_color="black",
                state="disabled"  # 🔒 Prevent manual editing
            )


        elif field == "Cash With":
            widget = ctk.CTkComboBox(
                other_frame,
                width=200,
                values=["COUNTER", "IOB-ACCOUNT", "TEKNIX", "BETA-ACCOUNT", "CUSTOMER", "OTHERS"],
                fg_color="white",
                text_color="black"
            )
            widget.set("COUNTER")

        elif field == "Collection Type":
            widget = ctk.CTkComboBox(
                other_frame,
                width=200,
                values=["ADAPTOR","MODEM","ONT","CABLING","RECONNECTION","SHIFTING","MAINTENANCE","PENALTY","OTHERS"],
                fg_color="white",
                text_color="black"
            )
            widget.set("ADAPTOR")  # default value


        else:
            widget = ctk.CTkEntry(
                other_frame,
                width=200,
                fg_color="white",
                text_color="black"
            )

        widget.place(x=180, y=y_pos)
        entries[field] = widget
        y_pos += 50

    # ---------- AUTO BALANCE CALCULATION ----------
    def calculate_balance(event=None):
        bill = entries["Bill Amount"].get().strip()
        cash = entries["Cash Received"].get().strip()

        try:
            bill_value = float(bill) if bill else 0
            cash_value = float(cash) if cash else 0
            balance = bill_value - cash_value
        except ValueError:
            balance = 0

        entries["Balance"].configure(state="normal")
        entries["Balance"].delete(0, "end")
        entries["Balance"].insert(0, f"{balance:.2f}")

        # 🔴 Turn red if negative
        if balance < 0:
            entries["Balance"].configure(text_color="red")
        else:
            entries["Balance"].configure(text_color="black")

        entries["Balance"].configure(state="disabled")

    # ---------- VALIDATION FUNCTION ----------
    def validate_entries():
        for field, widget in entries.items():
            value = widget.get().strip()

            if field in ["FTTH No", "Contact No"]:
                if len(value) != 10:
                    check_btn.configure(state="disabled")
                    return

            elif field not in ["Cash With", "Collection Type", "Balance", "Date"] and value == "":
                check_btn.configure(state="disabled")
                return

        check_btn.configure(state="normal")

    # ---------- DATE PICKER ----------
    def open_calendar(event=None):
        cal_win = ctk.CTkToplevel(parent)
        cal_win.title("Select Date")
        cal_win.geometry("300x320")
        cal_win.grab_set()

        cal = Calendar(
            cal_win,
            selectmode="day",
            date_pattern="dd-mm-yyyy"
        )
        cal.pack(pady=20)

        def select_date():
            selected = cal.get_date()
            entries["Date"].configure(state="normal")
            entries["Date"].delete(0, "end")
            entries["Date"].insert(0, selected)
            entries["Date"].configure(state="readonly")
            cal_win.destroy()

        ctk.CTkButton(
            cal_win,
            text="OK",
            command=select_date
        ).pack(pady=10)



    ### Create Excel Save Function
    def save_to_excel(data_dict):
        folder_path = "D:/BETA-SOFT"
        file_name = "beta_soft.xlsx"
        full_path = os.path.join(folder_path, file_name)
        os.makedirs(folder_path, exist_ok=True)

        # If file exists → load it
        if os.path.exists(full_path):
            workbook = load_workbook(full_path)
        else:
            workbook = Workbook()
        # If sheet "other" exists → use it
        if "other" in workbook.sheetnames:
            sheet = workbook["other"]
        else:
            sheet = workbook.create_sheet("other")
            # Write header only first time
            headers = ["Date", "Time"] + list(data_dict.keys())
            sheet.append(headers)

        now = datetime.now()
        date_str = "'"+now.strftime("%d-%m-%Y")
        time_str = now.strftime("%H:%M:%S")
        row_data = [date_str, time_str] + list(data_dict.values())
        sheet.append(row_data)
        workbook.save(full_path)

    # ---------- CHECK BUTTON ACTION ----------
    def show_verification():
        popup = ctk.CTkToplevel(parent)
        popup.title("OTHER COLLECTION VERIFICATION ")
        popup.geometry("450x620")
        popup.transient(parent)
        popup.grab_set()  # 🔒 MODAL
        popup.resizable(False, False)

        popup_frame = ctk.CTkFrame(popup)
        popup_frame.pack(fill="both", expand=True, padx=10, pady=10)

        popup_frame.columnconfigure(0, weight=1)
        popup_frame.columnconfigure(1, weight=2)

        title = ctk.CTkLabel(
            popup_frame,
            text="OTHER COLLECTION VERIFICATION DETAILS",
            font=("Arial", 18, "bold")
        )
        title.grid(row=0, column=0, columnspan=2, pady=(10, 20))

        row = 1
        for field, widget in entries.items():
            value = widget.get()

            ctk.CTkLabel(
                popup_frame,
                text=field,
                anchor="w"
            ).grid(row=row, column=0, padx=10, pady=6, sticky="w")

            ctk.CTkLabel(
                popup_frame,
                text=value,
                anchor="w"
            ).grid(row=row, column=1, padx=10, pady=6, sticky="w")

            row += 1

        # Buttons
        btn_frame = ctk.CTkFrame(popup_frame)
        btn_frame.grid(row=row, column=0, columnspan=2, pady=20)

        ctk.CTkButton(
            btn_frame,
            text="Back",
            width=100,
            command=popup.destroy
        ).pack(side="left", padx=10)

        def confirm_button_action():
            collected_data = {}
            for field, widget in entries.items():
                value = widget.get()
                if field == "Date":
                    value = f"'{value}"
                collected_data[field] = value
            print("Stored Data:", collected_data)
            save_to_excel(collected_data)
            popup.destroy()
            clear_entries()
            entries["FTTH No"].focus()

        ctk.CTkButton(
            btn_frame,
            text="Confirm",
            width=100,
            fg_color="green",
            command=confirm_button_action
        ).pack(side="left", padx=10)


    # ---------- Buttons ----------
    check_btn = ctk.CTkButton(
        other_frame,
        text="Check",
        width=100,
        fg_color="green",
        text_color="white",
        state="disabled",  # 🔒 initially disabled
        command=show_verification
    )
    check_btn.place(x=60, y=y_pos + 20)
    for field, widget in entries.items():
        if field != "Cash With":
            widget.bind("<KeyRelease>", lambda event: validate_entries())
    entries["Date"].bind("<Button-1>", open_calendar)

    # Bind auto calculation
    entries["Bill Amount"].bind("<KeyRelease>", calculate_balance)
    entries["Cash Received"].bind("<KeyRelease>", calculate_balance)


    # ---------- FULL ENTER KEY NAVIGATION ----------

    field_order = [
        "FTTH No",
        "Name",
        "Contact No",
        "Bill Amount",
        "Cash Received",
        "Remarks"
    ]

    def focus_next(event, current_field):
        if current_field in field_order:
            index = field_order.index(current_field)
            if index + 1 < len(field_order):
                next_field = field_order[index + 1]
                entries[next_field].focus()
            else:
                check_btn.focus()  # Last field → move to Check button

    # Bind Enter key for each field
    for field in field_order:
        entries[field].bind(
            "<Return>",
            lambda event, f=field: focus_next(event, f)
        )

    # Press Enter on Check button to trigger it
    check_btn.bind("<Return>", lambda event: check_btn.invoke())

    ####################

    def clear_entries():
        for field, widget in entries.items():
            # ✅ Reset Date to today
            if field == "Date":
                widget.configure(state="normal")
                widget.delete(0, "end")
                widget.insert(0, datetime.now().strftime("%d-%m-%Y"))
                widget.configure(state="readonly")

            # Enable balance temporarily to clear
            elif field == "Balance":
                widget.configure(state="normal")
                widget.delete(0, "end")
                widget.configure(state="disabled")
            # Reset combobox
            elif field == "Cash With":
                widget.set("COUNTER")
            elif field == "Collection Type":
                widget.set("INSTALLATION")

            # Normal entries
            else:
                widget.delete(0, "end")

        validate_entries()

    cancel_btn = ctk.CTkButton(
        other_frame,
        text="Cancel",
        width=100,
        fg_color="red",
        text_color="white",
        command=clear_entries  # clears instead of destroying
    )
    cancel_btn.place(x=200, y=y_pos + 20)

    ####################

    return other_frame

